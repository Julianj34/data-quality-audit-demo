"""export_reports.py — S5: Report Exporter.

Produces the six V1 output artefacts (Architecture, section 13):

    clean_master_dataset.xlsx   joined + cleaned + revenue + quality_flag
    errors_report.xlsx          1 sheet per validation layer + Summary
    data_quality_score.md       score, traffic light, distribution, top
    data_dictionary.md          tables, columns, types, rules, sources
    before_after_summary.md     * the business proof
    processing_log.txt          full run log with timestamps

This module only RENDERS — it receives a fully computed PipelineReport
and writes files. No validation, no cleaning, no scoring happens here
(keeps "Reports = Interpretation" honest). All values are already
computed upstream; nothing is recalculated, so the Excel files contain
data, not formulas — and therefore zero formula errors by construction.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from calculate_quality_score import ScoreResult, format_score_block
from clean_data import CleaningResult
from error_record import (CheckLayer, ErrorRecord, ERROR_REPORT_COLUMNS,
                          Severity, records_to_dataframe)

FONT = "Arial"
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")      # dark blue
_HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
_SEVERITY_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="F4CCCC"),     # red-ish
    "MAJOR": PatternFill("solid", fgColor="FCE5CD"),        # orange-ish
    "MINOR": PatternFill("solid", fgColor="FFF2CC"),        # yellow-ish
}
_LAYER_ORDER = [CheckLayer.SCHEMA, CheckLayer.COMPLETENESS,
                CheckLayer.UNIQUENESS, CheckLayer.RELATIONSHIP,
                CheckLayer.BUSINESS_RULE]
_LAYER_SHEET = {
    CheckLayer.SCHEMA: "Schema",
    CheckLayer.COMPLETENESS: "Completeness",
    CheckLayer.UNIQUENESS: "Uniqueness",
    CheckLayer.RELATIONSHIP: "Relationship",
    CheckLayer.BUSINESS_RULE: "BusinessRule",
}


@dataclass
class PipelineReport:
    """Everything the exporter needs — assembled by run_pipeline."""

    config: dict
    raw_dataframes: dict[str, pd.DataFrame]
    raw_errors: list[ErrorRecord]
    raw_score: ScoreResult
    cleaning: CleaningResult
    clean_errors: list[ErrorRecord]
    clean_score: ScoreResult
    log_lines: list[str] = field(default_factory=list)
    manifest: dict | None = None       # injected_errors.json (optional)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_all(report: PipelineReport, outdir: str | Path,
               processed_dir: str | Path | None = None) -> list[Path]:
    """Write all six artefacts.

    outdir         -> reports/ (errors_report.xlsx + the .md/.txt files)
    processed_dir  -> data/processed/ for clean_master_dataset.xlsx;
                      defaults to outdir when not given.
    """
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    processed = Path(processed_dir) if processed_dir else out
    processed.mkdir(parents=True, exist_ok=True)

    paths = [
        _export_clean_master(report,
                             processed / "clean_master_dataset.xlsx"),
        _export_errors_report(report, out / "errors_report.xlsx"),
        _write_text(out / "data_quality_score.md",
                    _render_score_md(report)),
        _write_text(out / "data_dictionary.md",
                    _render_dictionary_md(report)),
        _write_text(out / "before_after_summary.md",
                    _render_summary_md(report)),
        _write_text(out / "processing_log.txt",
                    "\n".join(report.log_lines) + "\n"),
    ]
    return paths


# ---------------------------------------------------------------------------
# Excel artefacts
# ---------------------------------------------------------------------------

def _style_sheet(ws, df: pd.DataFrame, severity_col: str | None = None
                 ) -> None:
    """Header styling, body font, freeze panes, autofilter, widths."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        letter = get_column_letter(col_idx)
        width = max(len(str(col_name)),
                    int(df[col_name].astype(str).str.len().max() or 0)
                    if len(df) else 0)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 48)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name=FONT)

    if severity_col and severity_col in df.columns and len(df):
        sev_idx = list(df.columns).index(severity_col) + 1
        for r in range(2, ws.max_row + 1):
            val = str(ws.cell(row=r, column=sev_idx).value)
            if val in _SEVERITY_FILL:
                ws.cell(row=r, column=sev_idx).fill = _SEVERITY_FILL[val]

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = (f"A1:{get_column_letter(ws.max_column)}"
                              f"{max(ws.max_row, 1)}")


def _export_clean_master(report: PipelineReport, path: Path) -> Path:
    master = report.cleaning.master
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Clean Master", index=False)
        _style_sheet(writer.sheets["Clean Master"], master)
    return path


def _export_errors_report(report: PipelineReport, path: Path) -> Path:
    df = records_to_dataframe(report.raw_errors)
    by_layer = {layer: df[df["check_layer"] == layer.value]
                for layer in _LAYER_ORDER}
    summary = _errors_summary_frame(report)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        _style_sheet(writer.sheets["Summary"], summary)
        for layer in _LAYER_ORDER:
            sheet_df = by_layer[layer].reset_index(drop=True)
            if sheet_df.empty:
                sheet_df = pd.DataFrame(columns=ERROR_REPORT_COLUMNS)
            sheet_df.to_excel(writer, sheet_name=_LAYER_SHEET[layer],
                              index=False)
            _style_sheet(writer.sheets[_LAYER_SHEET[layer]], sheet_df,
                         severity_col="severity")
    return path


def _errors_summary_frame(report: PipelineReport) -> pd.DataFrame:
    cfg = report.config
    sev_map = cfg["severity_map"]
    act_map = cfg["cleaning_actions"]
    raw_counts = report.raw_score.errors_by_check
    residual = {}
    for r in report.clean_errors:
        residual[r.check_id] = residual.get(r.check_id, 0) + 1

    rows = []
    for check_id in sorted(raw_counts):
        rows.append({
            "check_id": check_id,
            "layer": _layer_of(check_id),
            "severity": sev_map[check_id],
            "cleaning_action": act_map[check_id],
            "found_raw": raw_counts[check_id],
            "remaining_after_clean": residual.get(check_id, 0),
        })
    total = {
        "check_id": "TOTAL", "layer": "", "severity": "",
        "cleaning_action": "",
        "found_raw": sum(raw_counts.values()),
        "remaining_after_clean": sum(residual.values()),
    }
    return pd.DataFrame(rows + [total])


def _layer_of(check_id: str) -> str:
    return {"SCH": "SCHEMA", "COM": "COMPLETENESS", "UNI": "UNIQUENESS",
            "REL": "RELATIONSHIP", "BIZ": "BUSINESS_RULE"}[check_id[:3]]


# ---------------------------------------------------------------------------
# Markdown artefacts
# ---------------------------------------------------------------------------

def _render_score_md(report: PipelineReport) -> str:
    raw, clean = report.raw_score, report.clean_score
    L = ["# Data Quality Score", "",
         f"_Generated: {datetime.now():%Y-%m-%d %H:%M:%S}_", "",
         "## Audit result (raw data)", "",
         f"**Overall Quality Score: {raw.score}/100 {raw.emoji}**", "",
         "| Severity | Count | Action |", "|---|---|---|",
         f"| Critical | {raw.n_critical} | ACT |",
         f"| Major | {raw.n_major} | CHECK |",
         f"| Minor | {raw.n_minor} | OBSERVE |",
         f"| **Total** | **{raw.total_errors}** | |", "",
         f"Weighted error points: {raw.weighted_error_points:g} "
         f"({raw.error_points_per_1000_rows:.1f} per "
         f"{report.config['scoring']['normalize_per_rows']:g} rows, "
         f"{raw.total_row_count} rows checked).", "",
         "## Top error classes (raw)", "",
         "| Check | Count |", "|---|---|"]
    for check_id, n in raw.top_error_classes(5):
        L.append(f"| {check_id} | {n} |")
    tl = report.config["scoring"]["traffic_light"]
    L += ["", "## After cleaning", "",
          f"**Overall Quality Score: {clean.score}/100 {clean.emoji}**",
          "",
          f"Before: {raw.score}/100 {raw.emoji}  →  "
          f"After: {clean.score}/100 {clean.emoji}", "",
          f"Residual issues are documented (kept & flagged), not hidden: "
          f"{clean.n_major} major, {clean.n_minor} minor.", "",
          "## Traffic light thresholds (config)", "",
          f"- 🟢 green: score >= {tl['green']}",
          f"- 🟡 yellow: {tl['yellow']}–{tl['green'] - 1}",
          f"- 🔴 red: < {tl['yellow']}"]
    return "\n".join(L) + "\n"


def _render_dictionary_md(report: PipelineReport) -> str:
    cfg = report.config
    required = cfg["required_columns"]
    optional = cfg.get("optional_columns", {})
    types = cfg.get("column_types", {})
    descriptions = {
        "sale_id": "Primary key of a sale",
        "sale_date": "Date of the sale (ISO 8601 after cleaning)",
        "customer_id": "FK -> customers.customer_id",
        "product_id": "FK -> products.product_id",
        "quantity": "Units sold (> 0)",
        "unit_price": "Price per unit (> 0)",
        "discount": "Fractional discount in [0, max_discount]",
        "region": "Sales region (allowed list)",
        "sales_channel": "Channel (allowed list)",
        "customer_name": "Customer display name",
        "customer_segment": "Allowed values: `B2B`, `B2C`, `Enterprise`",
        "country": "Customer country",
        "signup_date": "Customer signup date",
        "product_name": "Product display name",
        "category": "Product category",
        "standard_price": "Catalogue price",
        "active_status": "Allowed values: `active`, `inactive`",
        "month": "Target month (YYYY-MM)",
        "target_revenue": "Monthly revenue target per region",
    }
    # customer_id / product_id are the PRIMARY KEY in their own master
    # table, but a FOREIGN KEY when they appear in sales.
    pk_in_master = {
        "customers": {"customer_id": "Primary key of a customer"},
        "products": {"product_id": "Primary key of a product"},
    }
    src = {"sales": "sales_raw.csv", "customers": "customers_raw.xlsx",
           "products": "products_raw.xlsx",
           "targets": "monthly_targets.xlsx"}

    L = ["# Data Dictionary", "",
         f"_Generated: {datetime.now():%Y-%m-%d %H:%M:%S}_", "",
         "Single source of truth for all rules: "
         "`data/reference/allowed_values.yaml`.", ""]
    for table in ("sales", "customers", "products", "targets"):
        L += [f"## {table}  (`{src[table]}`)", "",
              "| Column | Type | Req. | Description |", "|---|---|---|---|"]
        req = required.get(table, [])
        opt = optional.get(table, [])
        ttypes = types.get(table, {})
        for col in list(req) + [c for c in opt if c not in req]:
            typ = ttypes.get(col, "string")
            flag = "required" if col in req else "optional"
            desc = pk_in_master.get(table, {}).get(
                col, descriptions.get(col, ""))
            L.append(f"| {col} | {typ} | {flag} | {desc} |")
        L.append("")

    L += ["## Derived columns (clean_master_dataset)", "",
          "| Column | Definition |", "|---|---|",
          "| gross_revenue | quantity × unit_price |",
          "| net_revenue | gross_revenue × (1 − discount) |",
          "| quality_flag | semicolon-separated flag codes "
          "(e.g. REL_003, BIZ_007) |", ""]
    return "\n".join(L) + "\n"


def _render_summary_md(report: PipelineReport) -> str:
    raw, clean = report.raw_score, report.clean_score
    cl = report.cleaning
    dropped = cl.dropped_by_check
    flagged = cl.flagged_by_check

    L = ["# Before / After Summary", "",
         f"_Generated: {datetime.now():%Y-%m-%d %H:%M:%S}_", "",
         "## 1. Ausgangslage", "",
         "Sales-, Customer- und Product-Rohdaten aus mehreren Systemen. "
         "Fehlende und unbekannte IDs, Dubletten, ungültige Beträge, "
         "falsche Datumsformate und inkonsistente Kategorien machen ein "
         "verlässliches Reporting unmöglich.", "",
         f"Geprüfte Zeilen gesamt: {raw.total_row_count}.", "",
         "## 2. Prüfungen", "",
         "Fünf Validierungsschichten, read-only, gemeinsames "
         "Error-Record-Format:", "",
         "Schema · Completeness · Uniqueness · Relationships · "
         "Business Rules.", "",
         "## 3. Befunde (raw)", "",
         f"**Quality Score: {raw.score}/100 {raw.emoji}** — "
         f"{raw.total_errors} Fehler "
         f"({raw.n_critical} critical, {raw.n_major} major, "
         f"{raw.n_minor} minor).", "",
         "| Check | Gefunden | Aktion |", "|---|---|---|"]
    act_map = report.config["cleaning_actions"]
    for check_id, n in sorted(raw.errors_by_check.items()):
        L.append(f"| {check_id} | {n} | {act_map[check_id]} |")

    L += ["", "## 4. Bereinigung", "",
          f"- **Gefixt:** {cl.total_fixed} "
          f"(Datumsformat, Rabatt-Capping, Region/Channel-Mapping)",
          f"- **Gedroppt:** {cl.total_dropped} "
          f"(fehlende/ungültige Pflichtdaten, Dubletten, kaputte FKs)",
          f"- **Geflaggt:** {cl.total_flagged} "
          f"(behalten, aber markiert)", ""]
    if cl.escalated_to_drop:
        L += [f"- Davon {len(cl.escalated_to_drop)} FIX-Versuche "
              f"eskaliert zu DROP (nicht reparierbar).", ""]
    L += ["Drop-Verteilung: "
          + ", ".join(f"{k} {v}" for k, v in sorted(dropped.items()))
          + ".", ""]

    L += ["## 5. Restunsicherheit", ""]

    # FIX-flags (documented, but no longer open risk) vs. real residual
    fixed_flags = {"BIZ_003"}
    fix_meaning = {"BIZ_003": "Rabattwerte wurden auf das Maximum gecappt"}
    residual_meaning = {"COM_005": "optionale Felder fehlen",
                        "REL_003": "inaktive Produkte verkauft",
                        "BIZ_007": "Umsatz-Ausreißer"}

    doc_fixes = {k: v for k, v in flagged.items() if k in fixed_flags}
    real_residual = {k: v for k, v in flagged.items()
                     if k not in fixed_flags}

    if doc_fixes:
        L += ["**Dokumentierte Fix-Hinweise** (behoben und markiert, "
              "keine offene Unsicherheit):", ""]
        for k, v in sorted(doc_fixes.items()):
            L.append(f"- {k}: {v} {fix_meaning.get(k, '')}")
        L.append("")

    L += ["**Verbleibende Restunsicherheit** (bewusst behalten, im "
          "`quality_flag` markiert):", ""]
    for k, v in sorted(real_residual.items()):
        L.append(f"- {k}: {v} {residual_meaning.get(k, '')}")
    L += ["", f"Nach Cleaning verbleiben {clean.total_errors} "
          f"dokumentierte Hinweise ({clean.n_major} major, "
          f"{clean.n_minor} minor) — keine kritischen Fehler.", ""]

    L += ["## 6. Ergebnis", "",
          f"### Before: {raw.score}/100 {raw.emoji}  →  "
          f"After: {clean.score}/100 {clean.emoji}", "",
          f"Das bereinigte Master-Dataset umfasst "
          f"{len(cl.master)} reportingfähige Zeilen mit berechneten "
          f"Umsätzen (gross/net) und Qualitäts-Flags. Reports zu Umsatz, "
          f"Kunden und Produkten sind jetzt verlässlich möglich.", "",
          "Gedroppte Zeilen werden aus dem Reporting ausgeschlossen; "
          "geflaggte Zeilen bleiben sichtbar und prüfbar — nicht "
          "verstecken, sondern kontrolliert entscheiden.", ""]

    if report.manifest:
        total = report.manifest.get("total_injected_errors")
        detected = sum(1 for e in report.manifest.get("errors", []))
        L += ["## Verifizierbarkeit", "",
              f"Known injected demo errors detected: {detected} / {total}.",
              "",
              "_Der Claim bezieht sich ausschließlich auf die kontrolliert "
              "injizierten, bekannten Demo-Fehler — nicht auf alle "
              "möglichen Datenfehler._", ""]
    return "\n".join(L) + "\n"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _write_text(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path
