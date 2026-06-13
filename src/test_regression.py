"""test_regression.py — Full regression suite for V1 steps 1-3.

Self-contained: generates its own demo data into temp directories,
never touches data/raw. Run from anywhere:

    python test_regression.py

Covers three layers of confidence:
    POSITIVE   the system does what it claims (142/142, reproducible)
    NEGATIVE   the system fails loudly when it should (aborts, errors)
    META       the PROOF itself is trustworthy (the verifier detects
               both missed and phantom errors — it is not a rubber
               stamp that always prints green)

Exit code 0 only if every test passes.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import sys
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT / "src"))

import pandas as pd
import yaml

from error_record import (ConfigError, ErrorFactory, KNOWN_CHECKS,
                          validate_config)
from load_data import (DataLoadError, is_iso_date, load_raw_data,
                       try_parse_date, try_parse_float)
import validate_schema
import validate_completeness
import validate_uniqueness
import validate_relationships
import validate_business_rules

CONFIG_PATH = REPO_ROOT / "data" / "reference" / "allowed_values.yaml"
GENERATOR = REPO_ROOT / "src" / "generate_demo_data.py"
VERIFIER = REPO_ROOT / "verify_detection.py"

VALIDATOR_MODULES = [validate_schema, validate_completeness,
                     validate_uniqueness, validate_relationships,
                     validate_business_rules]

_results: list[tuple[str, bool, str]] = []


def report(test_id: str, name: str, ok: bool, detail: str = "") -> None:
    _results.append((test_id, ok, name))
    status = "PASS" if ok else "FAIL"
    line = f"  [{status}] {test_id}  {name}"
    if detail:
        line += f" — {detail}"
    print(line)


def md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def load_config() -> dict:
    return yaml.safe_load(CONFIG_PATH.read_text())


def run_detection(raw_dir: Path) -> list:
    """Run S1 + S2 in-process, return the ErrorRecord list."""
    config = load_config()
    factory = ErrorFactory(config)
    dataframes = load_raw_data(raw_dir)
    validate_schema.assert_file_level_schema(dataframes, config)
    detected = []
    for module in VALIDATOR_MODULES:
        detected.extend(module.run_checks(dataframes, config, factory))
    return detected


def triples(records) -> Counter:
    return Counter((r.table, r.row_ref, r.check_id) for r in records)


# ===========================================================================

def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="dqa_regression_"))
    run_a, run_b = tmp / "run_a", tmp / "run_b"

    print(f"Regression suite — temp workspace: {tmp}\n")

    # --- T01: config fulfils the Clean Core contract -----------------------
    try:
        config = load_config()
        validate_config(config)
        complete = (set(config["severity_map"]) == set(KNOWN_CHECKS)
                    and set(config["cleaning_actions"]) == set(KNOWN_CHECKS))
        report("T01", "Config contract (validate_config + completeness)",
               complete, f"{len(KNOWN_CHECKS)} checks mapped")
    except Exception as e:
        report("T01", "Config contract", False, str(e))

    # --- T02: clean core unit tests -----------------------------------------
    r = subprocess.run([sys.executable,
                        str(REPO_ROOT / "test_clean_core.py")],
                       capture_output=True, text=True)
    report("T02", "Clean core unit tests (test_clean_core.py)",
           r.returncode == 0 and "all contract tests passed" in r.stdout)

    # --- T03: generator runs, self-check passes -----------------------------
    r = subprocess.run([sys.executable, str(GENERATOR),
                        "--outdir", str(run_a)],
                       capture_output=True, text=True)
    manifest_a = json.loads((run_a / "injected_errors.json").read_text()) \
        if r.returncode == 0 else {}
    ok = (r.returncode == 0
          and manifest_a.get("total_injected_errors") == 142
          and len(manifest_a.get("errors", [])) == 142)
    report("T03", "Generator + self-check", ok,
           f"{manifest_a.get('total_injected_errors')} errors, "
           f"{len(manifest_a.get('counts_by_class', {}))} classes")

    # --- T04: byte-identical reproducibility --------------------------------
    subprocess.run([sys.executable, str(GENERATOR),
                    "--outdir", str(run_b)],
                   capture_output=True, text=True)
    csv_same = md5(run_a / "sales_raw.csv") == md5(run_b / "sales_raw.csv")
    manifest_same = (json.loads((run_a / "injected_errors.json").read_text())
                     == json.loads((run_b / "injected_errors.json")
                                   .read_text()))
    xlsx_same = all(
        pd.read_excel(run_a / f).equals(pd.read_excel(run_b / f))
        for f in ("customers_raw.xlsx", "products_raw.xlsx",
                  "monthly_targets.xlsx"))
    report("T04", "Reproducibility (seed 42, two runs identical)",
           csv_same and manifest_same and xlsx_same)

    # --- T05: verifier 142/142, exit 0 ---------------------------------------
    r = subprocess.run([sys.executable, str(VERIFIER),
                        "--raw-dir", str(run_a),
                        "--config", str(CONFIG_PATH)],
                       capture_output=True, text=True)
    report("T05", "Detection vs manifest (verify_detection.py)",
           r.returncode == 0 and "142 / 142" in r.stdout
           and "VERIFIED" in r.stdout)

    # --- T06: detection is deterministic -------------------------------------
    d1, d2 = run_detection(run_a), run_detection(run_a)
    report("T06", "Detection determinism (two in-process runs)",
           triples(d1) == triples(d2) and len(d1) == 142,
           f"{len(d1)} == {len(d2)}")

    # --- T07: zero false positives on the clean baseline ---------------------
    # Remove every manifest-affected sales row; dedupe masters (keep first,
    # NOT remove — clean sales legitimately reference those keys).
    dfs = load_raw_data(run_a)
    sales_refs = {e["row_ref"] for e in manifest_a["errors"]
                  if e["table"] == "sales"}
    clean_dfs = {
        "sales": dfs["sales"][~dfs["sales"]["sale_id"].isin(sales_refs)],
        "customers": dfs["customers"].drop_duplicates("customer_id"),
        "products": dfs["products"].drop_duplicates("product_id"),
        "targets": dfs["targets"],
    }
    factory = ErrorFactory(load_config())
    found = []
    for module in VALIDATOR_MODULES:
        found.extend(module.run_checks(clean_dfs, load_config(), factory))
    report("T07", "Clean baseline produces ZERO errors",
           len(found) == 0,
           f"{len(clean_dfs['sales'])} clean rows, {len(found)} errors")

    # --- T08: META — verifier catches a MISSED error --------------------------
    # Repair one injected error in the data; manifest now expects an error
    # the data no longer contains -> 1 missed, exit 1.
    tampered = tmp / "tamper_fix"
    shutil.copytree(run_a, tampered)
    rel = next(e for e in manifest_a["errors"]
               if e["error_class"] == "REL_001")
    csv_text = (tampered / "sales_raw.csv").read_text()
    bad_id = re.search(r"customer_id (C-\d+)", rel["detail"]).group(1)
    (tampered / "sales_raw.csv").write_text(
        csv_text.replace(bad_id, "C-0001", 1))
    r = subprocess.run([sys.executable, str(VERIFIER),
                        "--raw-dir", str(tampered),
                        "--config", str(CONFIG_PATH)],
                       capture_output=True, text=True)
    report("T08", "META: verifier catches missed error (repaired row)",
           r.returncode == 1 and "Missed injections:     1" in r.stdout)

    # --- T09: META — verifier catches an UNEXPECTED detection ----------------
    # Break one clean row; data now contains an error the manifest doesn't
    # know -> 1 unexpected, exit 1.
    tampered2 = tmp / "tamper_add"
    shutil.copytree(run_a, tampered2)
    df = pd.read_csv(tampered2 / "sales_raw.csv", dtype=str,
                     keep_default_na=False)
    clean_idx = df.index[~df["sale_id"].isin(sales_refs)][0]
    df.loc[clean_idx, "region"] = "Narnia"
    df.to_csv(tampered2 / "sales_raw.csv", index=False)
    r = subprocess.run([sys.executable, str(VERIFIER),
                        "--raw-dir", str(tampered2),
                        "--config", str(CONFIG_PATH)],
                       capture_output=True, text=True)
    report("T09", "META: verifier catches unexpected detection",
           r.returncode == 1 and "Unexpected detections: 1" in r.stdout)

    # --- T10: SCH_001 file-level abort ----------------------------------------
    dfs10 = load_raw_data(run_a)
    dfs10["sales"] = dfs10["sales"].drop(columns=["customer_id"])
    try:
        validate_schema.assert_file_level_schema(dfs10, load_config())
        report("T10", "SCH_001 abort on missing column", False,
               "no abort raised")
    except validate_schema.SchemaAbortError:
        report("T10", "SCH_001 abort on missing column", True)

    # --- T11: missing input file aborts loading --------------------------------
    partial = tmp / "partial"
    shutil.copytree(run_a, partial)
    (partial / "products_raw.xlsx").unlink()
    try:
        load_raw_data(partial)
        report("T11", "Missing input file -> DataLoadError", False)
    except DataLoadError as e:
        report("T11", "Missing input file -> DataLoadError",
               "products_raw.xlsx" in str(e))

    # --- T12: broken config fails at startup -----------------------------------
    broken = load_config()
    del broken["cleaning_actions"]["BIZ_004"]
    try:
        ErrorFactory(broken)
        report("T12", "Broken config -> ConfigError at startup", False)
    except ConfigError:
        report("T12", "Broken config -> ConfigError at startup", True)

    # --- T13: parsing helper edge cases ------------------------------------------
    cases_ok = all([
        is_iso_date("2024-01-05") is True,
        is_iso_date("2024-1-5") is False,        # strptime leniency closed
        is_iso_date("2024-02-30") is False,      # invalid calendar date
        is_iso_date(" 2024-01-05") is False,     # whitespace = format error
        try_parse_date("15.03.2024") is not None,
        try_parse_date("2024-02-30") is None,
        try_parse_date("garbage") is None,
        try_parse_float("69.29") == 69.29,
        try_parse_float("-5") == -5.0,
        try_parse_float("1,5") is None,          # decimal comma -> SCH_002
        try_parse_float(None) is None,
    ])
    report("T13", "Parsing helpers (strict ISO, calendar, garbage)",
           cases_ok)

    # --- T14: static cross-check code <-> registry --------------------------------
    # Match quoted check-ID literals anywhere (incl. mapping tables like
    # _MANDATORY_SALES / _KEYS) — not only check_id="..." keyword args.
    produced = set()
    id_pattern = re.compile(r'"((?:SCH|COM|UNI|REL|BIZ)_\d{3})"')
    for f in (REPO_ROOT / "src").glob("validate_*.py"):
        produced |= set(id_pattern.findall(f.read_text()))
    # SCH_001 aborts instead of producing a record — by design
    report("T14", "Every emitted check_id is registered "
                  "(KNOWN_CHECKS minus abort-only SCH_001)",
           produced == set(KNOWN_CHECKS) - {"SCH_001"},
           f"{len(produced)} check_ids in code")

    # --- T15: severity / cleaning_action wiring -------------------------------------
    by_check = {r.check_id: r for r in d1}
    expectations = [
        ("REL_001", "CRITICAL", "DROP"),
        ("BIZ_003", "MAJOR", "FIX"),
        ("COM_005", "MINOR", "FLAG"),
        ("BIZ_007", "MINOR", "FLAG"),
        ("UNI_001", "MAJOR", "DROP"),
    ]
    wiring_ok = all(
        by_check[c].severity.value == sev
        and by_check[c].cleaning_action.value == act
        for c, sev, act in expectations)
    report("T15", "Severity/action wired from config (spot checks)",
           wiring_ok)

    # --- T16: score formula on a hand-calculated case --------------------------------
    # 2 CRITICAL + 3 MAJOR + 4 MINOR on 1000 rows:
    # weighted = 2*1.0 + 3*0.4 + 4*0.1 = 3.6 -> 100-3.6 = 96.4 -> 96, green
    from calculate_quality_score import (calculate_quality_score,
                                         count_total_rows,
                                         format_score_block,
                                         round_half_up)

    def make_records(spec):
        """spec: list of (check_id, n) — severities come from config."""
        fac = ErrorFactory(load_config())
        recs = []
        for check_id, n in spec:
            for i in range(n):
                recs.append(fac.create(
                    check_id=check_id, table="sales", row_ref=f"T-{i}",
                    column=None, invalid_value="x", rule="test",
                    message="synthetic"))
        return recs

    recs = make_records([("REL_001", 2),    # CRITICAL
                         ("UNI_001", 3),    # MAJOR
                         ("COM_005", 4)])   # MINOR
    res = calculate_quality_score(recs, 1000, load_config())
    report("T16", "Score formula matches hand calculation",
           (round(res.weighted_error_points, 1) == 3.6 and res.score == 96
            and res.traffic_light == "green"
            and res.n_critical == 2 and res.n_major == 3
            and res.n_minor == 4),
           f"score {res.score}, {res.traffic_light}")

    # --- T17: score on the real detection run (regression anchor) ---------------------
    # Severity distribution of the 142 injected errors is fixed by the
    # generator: 58 CRITICAL / 66 MAJOR / 18 MINOR.
    # weighted = 58*1.0 + 66*0.4 + 18*0.1 = 58 + 26.4 + 1.8 = 86.2
    # rows = 2006 + 204 + 53 + 60 = 2323 -> 37.1 pts/1000 -> 100-37.1=62.9
    # half-up -> score 63, red
    rows_total = count_total_rows(load_raw_data(run_a))
    res17 = calculate_quality_score(d1, rows_total, load_config())
    report("T17", "Real-run score (anchors severity distribution)",
           (res17.n_critical == 58 and res17.n_major == 66
            and res17.n_minor == 18
            and round(res17.weighted_error_points, 1) == 86.2
            and rows_total == 2323
            and res17.score == 63 and res17.traffic_light == "red"),
           f"58/66/18, weighted 86.2, rows {rows_total}, "
           f"score {res17.score}")

    # --- T18: edge cases (empty, clamp, rounding, zero rows) ---------------------------
    res_empty = calculate_quality_score([], 1000, load_config())
    # clamp: 200 CRITICAL on 100 rows -> 2000 pts/1000 -> 100-2000 < 0 -> 0
    res_clamp = calculate_quality_score(
        make_records([("REL_001", 200)]), 100, load_config())
    # rounding: 95 MINOR on 1000 rows -> 9.5 pts -> 90.5 -> half-up 91
    # (banker's round() would give 90 — this pins the half-up decision)
    res_round = calculate_quality_score(
        make_records([("COM_005", 95)]), 1000, load_config())
    try:
        calculate_quality_score([], 0, load_config())
        zero_rows_ok = False
    except ValueError:
        zero_rows_ok = True
    report("T18", "Edge cases: empty=100, clamp>=0, half-up, 0 rows",
           (res_empty.score == 100 and res_empty.traffic_light == "green"
            and res_clamp.score == 0
            and res_round.score == 91
            and round_half_up(90.5) == 91 and round_half_up(89.5) == 90
            and zero_rows_ok),
           f"clamp -> {res_clamp.score}, rounding 90.5 -> "
           f"{res_round.score}")

    # --- T19: traffic light thresholds from config ---------------------------------------
    # With MINOR weight 0.1 on 1000 rows the boundaries sit at:
    #  96 MINOR ->  9.6 pts -> 90.4 -> 90 -> green (boundary inclusive)
    # 106 MINOR -> 10.6 pts -> 89.4 -> 89 -> yellow
    # 296 MINOR -> 29.6 pts -> 70.4 -> 70 -> yellow (boundary inclusive)
    # 306 MINOR -> 30.6 pts -> 69.4 -> 69 -> red
    lights = [calculate_quality_score(make_records([("COM_005", n)]),
                                      1000, load_config())
              for n in (96, 106, 296, 306)]
    block = format_score_block(res17, load_config(), label="raw")
    report("T19", "Traffic light boundaries + decision block",
           ([(r.score, r.traffic_light) for r in lights]
            == [(90, "green"), (89, "yellow"),
                (70, "yellow"), (69, "red")]
            and "ACT" in block and "CHECK" in block
            and "OBSERVE" in block),
           "90/89/70/69 -> green/yellow/yellow/red")

    # =====================================================================
    # CLEANING (S4) — T20..T26
    # =====================================================================
    from clean_data import clean

    def detect(dfs):
        fac = ErrorFactory(load_config())
        validate_schema.assert_file_level_schema(dfs, load_config())
        recs = []
        for module in VALIDATOR_MODULES:
            recs.extend(module.run_checks(dfs, load_config(), fac))
        return recs

    raw_dfs = load_raw_data(run_a)
    raw_errors = detect(raw_dfs)
    cfg = load_config()
    res_clean = clean(raw_dfs, raw_errors, cfg)

    # --- T20: cleaning balance reconciles to 142 -------------------------
    # master-dedup (UNI_002+UNI_003) + sales-drops + real fixes + pure flags
    master_dedup = (res_clean.dropped_by_check.get("UNI_002", 0)
                    + res_clean.dropped_by_check.get("UNI_003", 0))
    sales_drops = res_clean.total_dropped - master_dedup
    pure_flags = (res_clean.flagged_by_check.get("COM_005", 0)
                  + res_clean.flagged_by_check.get("REL_003", 0)
                  + res_clean.flagged_by_check.get("BIZ_007", 0))
    balance = master_dedup + sales_drops + res_clean.total_fixed + pure_flags
    report("T20", "Cleaning balance reconciles to 142 injected",
           balance == 142 and master_dedup == 7 and sales_drops == 81
           and res_clean.total_fixed == 29 and pure_flags == 25,
           f"{master_dedup}+{sales_drops}+{res_clean.total_fixed}"
           f"+{pure_flags} = {balance}")

    # --- T21: before/after score (the business proof) --------------------
    raw_score = calculate_quality_score(
        raw_errors, count_total_rows(raw_dfs), cfg)
    clean_errors = detect(res_clean.cleaned)
    clean_score = calculate_quality_score(
        clean_errors, count_total_rows(res_clean.cleaned), cfg)
    report("T21", "Before/After score (63 red -> 98 green)",
           raw_score.score == 63 and raw_score.traffic_light == "red"
           and clean_score.score == 98
           and clean_score.traffic_light == "green",
           f"{raw_score.score} -> {clean_score.score}")

    # --- T22: residual errors are EXACTLY the FLAG classes ---------------
    residual = Counter(r.check_id for r in clean_errors)
    report("T22", "Residual errors == flagged classes (documented risk)",
           dict(residual) == {"COM_005": 13, "REL_003": 7, "BIZ_007": 5}
           and sum(residual.values()) == 25,
           f"{sum(residual.values())} residual: {dict(residual)}")

    # --- T23: FIX -> DROP escalation (unrepairable values) ---------------
    esc = Counter(e.check_id for e in res_clean.escalated_to_drop)
    esc_values = {e.original_value for e in res_clean.escalated_to_drop}
    report("T23", "FIX escalates to DROP when unrepairable",
           len(res_clean.escalated_to_drop) == 5
           and esc == {"BIZ_005": 3, "BIZ_006": 2}
           and {"Atlantis", "Mordor", "Springfield", "fax",
                "carrier-pigeon"} == esc_values,
           f"5 escalations: {dict(esc)}")

    # --- T24: master dataset integrity -----------------------------------
    m = res_clean.master
    flagged_rows = (m["quality_flag"].astype(str).str.len() > 0).sum()
    report("T24", "Master: no NaN revenue, no FK gaps, flags + revenue",
           (len(m) == 1925
            and int(m["gross_revenue"].isna().sum()) == 0
            and int(m["net_revenue"].isna().sum()) == 0
            and int(m["customer_name"].isna().sum()) == 0
            and int(m["product_name"].isna().sum()) == 0
            and flagged_rows == 35
            and "gross_revenue" in m.columns
            and "net_revenue" in m.columns),
           f"{len(m)} rows, {flagged_rows} flagged, 0 NaN/FK gaps")

    # --- T25: fix-log invariant + idempotence ----------------------------
    log_ok = all(e.original_value and e.fixed_value
                 and e.original_value != e.fixed_value
                 for e in res_clean.fix_log)
    # re-clean the cleaned data: nothing left to drop or fix, only flags
    res2 = clean(res_clean.cleaned, clean_errors, cfg)
    idempotent = (res2.total_dropped == 0 and res2.total_fixed == 0
                  and res2.total_flagged == 25
                  and len(res2.cleaned["sales"]) == 1925)
    report("T25", "Fix-log records original!=fixed; cleaning idempotent",
           log_ok and idempotent,
           f"{len(res_clean.fix_log)} fixes logged, re-clean stable")

    # --- T26: cleaning does not mutate its input -------------------------
    raw_after = load_raw_data(run_a)
    untouched = all(
        raw_dfs[t].reset_index(drop=True).equals(
            raw_after[t].reset_index(drop=True))
        for t in ("sales", "customers", "products", "targets"))
    report("T26", "clean() does not mutate input dataframes", untouched)

    # =====================================================================
    # EXPORT (S5) — T27..T31
    # =====================================================================
    import openpyxl
    from export_reports import PipelineReport, export_all

    manifest = json.loads((run_a / "injected_errors.json").read_text())
    pr = PipelineReport(
        config=cfg, raw_dataframes=raw_dfs, raw_errors=raw_errors,
        raw_score=raw_score, cleaning=res_clean,
        clean_errors=clean_errors, clean_score=clean_score,
        log_lines=[f"{datetime.now():%Y-%m-%d %H:%M:%S}  test run"],
        manifest=manifest)
    rep_dir = tmp / "reports"
    paths = export_all(pr, rep_dir)

    # --- T27: all six artefacts exist and are non-empty ------------------
    expected = {"clean_master_dataset.xlsx", "errors_report.xlsx",
                "data_quality_score.md", "data_dictionary.md",
                "before_after_summary.md", "processing_log.txt"}
    names = {p.name for p in paths}
    sizes_ok = all(p.stat().st_size > 0 for p in paths)
    report("T27", "All six artefacts written, non-empty",
           names == expected and sizes_ok and len(paths) == 6)

    # --- T28: excel structure + zero formula errors ----------------------
    wb_err = openpyxl.load_workbook(rep_dir / "errors_report.xlsx")
    wb_mas = openpyxl.load_workbook(rep_dir / "clean_master_dataset.xlsx")

    def formula_errors(wb):
        bad = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
        return sum(1 for ws in wb.worksheets for row in ws.iter_rows()
                   for c in row if isinstance(c.value, str)
                   and c.value in bad)

    report("T28", "Excel: 6 layer sheets + master, zero formula errors",
           (wb_err.sheetnames == ["Summary", "Schema", "Completeness",
                                  "Uniqueness", "Relationship",
                                  "BusinessRule"]
            and wb_mas.sheetnames == ["Clean Master"]
            and formula_errors(wb_err) == 0
            and formula_errors(wb_mas) == 0))

    # --- T29: markdown carries the key numbers ---------------------------
    summary_txt = (rep_dir / "before_after_summary.md").read_text()
    score_txt = (rep_dir / "data_quality_score.md").read_text()
    dict_txt = (rep_dir / "data_dictionary.md").read_text()
    report("T29", "Markdown carries 63->98, 142/142, derived columns",
           ("63/100" in summary_txt and "98/100" in summary_txt
            and "142 / 142" in summary_txt
            and "nicht auf alle" in summary_txt        # scope note
            and "63/100" in score_txt and "98/100" in score_txt
            and "gross_revenue" in dict_txt
            and "net_revenue" in dict_txt))

    # --- T30: errors_report Summary reconciles to 142 / 25 ---------------
    summary_df = pd.read_excel(rep_dir / "errors_report.xlsx",
                               sheet_name="Summary")
    total_row = summary_df[summary_df["check_id"] == "TOTAL"].iloc[0]
    report("T30", "errors_report Summary totals: 142 found, 25 remaining",
           (int(total_row["found_raw"]) == 142
            and int(total_row["remaining_after_clean"]) == 25),
           f"{int(total_row['found_raw'])} found, "
           f"{int(total_row['remaining_after_clean'])} remaining")

    # --- T31: master round-trips; revenue is internally consistent -------
    rt = pd.read_excel(rep_dir / "clean_master_dataset.xlsx")
    gross_ok = ((rt["quantity"] * rt["unit_price"]).round(2)
                == rt["gross_revenue"].round(2)).all()
    net_ok = ((rt["gross_revenue"] * (1 - rt["discount"].fillna(0)))
              .round(2) == rt["net_revenue"].round(2)).all()
    report("T31", "Master round-trips; gross/net revenue consistent",
           len(rt) == 1925 and gross_ok and net_ok
           and "quality_flag" in rt.columns,
           f"{len(rt)} rows, revenue checks pass")

    # =====================================================================
    # ORCHESTRATION (run_pipeline) — T32..T33
    # =====================================================================
    # --- T32: end-to-end run produces all artefacts in the right dirs ----
    pipe_reports = tmp / "pipe_reports"
    pipe_processed = tmp / "pipe_processed"
    r = subprocess.run(
        [sys.executable, str(REPO_ROOT / "run_pipeline.py"),
         "--raw-dir", str(run_a),
         "--config", str(CONFIG_PATH),
         "--reports-dir", str(pipe_reports),
         "--processed-dir", str(pipe_processed)],
        capture_output=True, text=True)
    master_here = (pipe_processed / "clean_master_dataset.xlsx").exists()
    reports_here = all((pipe_reports / f).exists() for f in (
        "errors_report.xlsx", "data_quality_score.md",
        "data_dictionary.md", "before_after_summary.md",
        "processing_log.txt"))
    log_has_scores = ("63/100" in r.stdout and "98/100" in r.stdout
                      and "142 / 142" in r.stdout)
    report("T32", "run_pipeline end-to-end: artefacts, dirs, log",
           (r.returncode == 0 and master_here and reports_here
            and log_has_scores),
           f"exit {r.returncode}, master in processed/, "
           f"5 files in reports/")

    # --- T33: pipeline fail paths return exit 1 with clear message -------
    r_missing = subprocess.run(
        [sys.executable, str(REPO_ROOT / "run_pipeline.py"),
         "--raw-dir", str(tmp / "does_not_exist"),
         "--config", str(CONFIG_PATH),
         "--reports-dir", str(tmp / "x"),
         "--processed-dir", str(tmp / "y")],
        capture_output=True, text=True)
    # missing column -> schema abort
    badschema = tmp / "badschema"
    shutil.copytree(run_a, badschema)
    bs = pd.read_csv(badschema / "sales_raw.csv", dtype=str,
                     keep_default_na=False)
    bs.drop(columns=["product_id"]).to_csv(
        badschema / "sales_raw.csv", index=False)
    r_schema = subprocess.run(
        [sys.executable, str(REPO_ROOT / "run_pipeline.py"),
         "--raw-dir", str(badschema), "--config", str(CONFIG_PATH),
         "--reports-dir", str(tmp / "x2"),
         "--processed-dir", str(tmp / "y2")],
        capture_output=True, text=True)
    report("T33", "Pipeline fail paths: exit 1 + clear abort message",
           (r_missing.returncode == 1
            and "ABORT (S1)" in r_missing.stderr
            and r_schema.returncode == 1
            and "ABORT (S2-A)" in r_schema.stderr
            and "product_id" in r_schema.stderr))

    # --- summary -------------------------------------------------------------------
    failed = [t for t in _results if not t[1]]
    print(f"\n{'=' * 60}")
    print(f"  {len(_results) - len(failed)} / {len(_results)} tests passed")
    if failed:
        for test_id, _, name in failed:
            print(f"  FAILED: {test_id} {name}")
        print("=" * 60)
        return 1
    print("  REGRESSION SUITE: ALL GREEN")
    print("=" * 60)
    shutil.rmtree(tmp, ignore_errors=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
